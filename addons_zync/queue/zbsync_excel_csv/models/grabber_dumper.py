from odoo import _, api, fields, models, SUPERUSER_ID
from odoo.exceptions import UserError, RedirectWarning, ValidationError
from odoo.tools.safe_eval import safe_eval
import io
from io import StringIO, BytesIO
import csv
import openpyxl

SEL_CONTENT_TYPES = [("csv", "CSV"), ("excel", "Excel")]

DEFAULT_MAX_COL_COUNT = 1000


class Mixin(models.AbstractModel):
    _name = "zbs.csvexcel.mixin"

    type = fields.Selection(SEL_CONTENT_TYPES, string="Type")
    first_line_headers = fields.Boolean("First Line Headers", default=True)
    fieldnames = fields.Char("Field names comma separated")
    csv_delimiter = fields.Char("CSV Delimiter", default=";")
    max_column_count = fields.Integer("Max Column Count")
    sample_file = fields.Binary("Sample File")

    def _get_stream(self, content):
        if isinstance(content, bytes):
            stream = BytesIO(content)
        elif isinstance(content, str):
            stream = StringIO(content)
        else:
            raise NotImplementedError(content)
        return stream

    def generate_mapper(self):
        if not self.sample_file:
            raise UserError(_("Please upload a sample file"))
        raise NotImplementedError("generate mapper")


class CSVGrabber(models.Model):
    _inherit = ["zbs.grabber", "zbs.csvexcel.mixin"]
    _name = "zbs.grabber.csvexcel"

    def process_record(self, instance, index, record, data):
        if not record:
            return

        if self.type == "csv":
            res = self._read_csv(record)
        elif self.type == "excel":
            res = self._read_excel(record)
        else:
            raise NotImplementedError(self.type)
        return res

    def _read_csv(self, content):
        stream = self._get_stream(content)
        csv_reader = csv.reader(stream, delimiter=self.csv_delimiter)
        records = []
        for i, row in enumerate(csv_reader):
            data = {}
            if not i:
                if self.first_line_headers:
                    fieldnames = list(row)
                else:
                    fieldnames = self.fieldnames.split(",")
            else:
                for i, fieldname in enumerate(fieldnames):
                    data[fieldname] = row[i]
                records.append(data)
        return records

    def _read_excel(self, content):
        stream = self._get_stream(content)
        wb = openpyxl.load_workbook(stream, data_only=True, read_only=True)
        ws = wb.active
        keys = []
        records = []
        max_cols = self.max_column_count or DEFAULT_MAX_COL_COUNT
        for idx, row in enumerate(
            ws.iter_rows(min_row=1, min_col=1, max_row=99999999, max_col=max_cols)
        ):
            if not idx and self.first_line_headers:
                keys = [x.value for x in row]
                continue
            record = {keys[i]: x.value for i, x in enumerate(row)}
            if any(record.values()):
                records.append(record)
        return records


class CSVDumper(models.Model):
    _inherit = ["zbs.dumper", "zbs.csvexcel.mixin"]
    _name = "zbs.dumper.csvexcel"

    def process(self, instance, data):
        if self.type == "csv":
            content = self._write_csv(data)
        elif self.type == "excel":
            content = self._write_excel(data)
        else:
            raise NotImplementedError(self.type)

        keep = self._keep_input(0, data, data)
        return instance.Result(
            {
                "content": content,
                "data": keep,
            }
        )

    def _write_csv(self, records):
        stream = io.StringIO()
        csv_writer = csv.writer(stream)
        if self.fieldnames:
            fieldnames = [x.strip() for x in self.fieldnames.split(",")]
        else:
            fieldnames = None
        for i, record in enumerate(records._iterate_records()):
            if not i and fieldnames is None:
                fieldnames = list(record.keys())

            if not i and self.first_line_headers:
                csv_writer.writerow(fieldnames)
            csv_writer.writerow([record[x] for x in fieldnames])
        stream.seek(0)
        res = stream.getvalue()
        return res

    def _write_excel(self, records):
        workbook = openpyxl.Workbook()

        # Get the active sheet in the workbook
        sheet = workbook.active

        row = 1
        fieldnames = None

        def _get_fieldnames(record):
            return (
                self.fieldnames.split(",") if self.fieldnames else list(record.keys())
            )

        first_record = next(records._iterate_records())
        if first_record and fieldnames is None:
            fieldnames = _get_fieldnames(first_record)

        if self.first_line_headers:
            for col_num, fieldname in enumerate(fieldnames, 1):
                cell = sheet.cell(row=row, column=col_num)
                cell.value = fieldname
            row += 1

        for row_num, row_data in enumerate(records._iterate_records(), row):
            for col_num, fieldname in enumerate(fieldnames, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = row_data[fieldname]

        # Save the workbook
        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream.getvalue()
